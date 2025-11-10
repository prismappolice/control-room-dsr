from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_login import login_required, current_user, login_user, logout_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from app.models import User, DSREntry, ControlRoomUpload, FORM_CONFIGS, DISTRICTS
from app import db, login_manager
import json
from datetime import datetime, date
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import io

# Blueprint definitions
main_bp = Blueprint('main', __name__)
auth_bp = Blueprint('auth', __name__)
admin_bp = Blueprint('admin', __name__)
district_bp = Blueprint('district', __name__)

# User loader for Flask-Login
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Main routes
@main_bp.route('/')
def index():
    return render_template('index.html')

# Authentication routes
@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user_type = request.form['user_type']
        
        user = User.query.filter_by(username=username, user_type=user_type).first()
        
        if user and user.check_password(password):
            login_user(user)
            if user.user_type == 'admin':
                return redirect(url_for('admin.dashboard'))
            elif user.user_type == 'district':
                return redirect(url_for('district.dashboard'))
            elif user.user_type == 'controlroom':
                return redirect(url_for('district.controlroom_dashboard'))
        else:
            flash('Invalid credentials', 'error')
    
    return render_template('login.html', districts=DISTRICTS)

@auth_bp.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('main.index'))

# Admin routes
@admin_bp.route('/dashboard')
@login_required
def dashboard():
    if current_user.user_type != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    # Get recent DSR entries
    recent_entries = DSREntry.query.order_by(DSREntry.updated_at.desc()).limit(10).all()
    
    # Get recent uploads
    recent_uploads = ControlRoomUpload.query.order_by(ControlRoomUpload.uploaded_at.desc()).limit(5).all()
    
    # Get statistics
    total_districts = 26  # Changed from len(DISTRICTS) to show 26 instead of 28
    total_forms = len(FORM_CONFIGS)
    today = date.today()
    today_entries = DSREntry.query.filter_by(date=today).count()
    
    return render_template('admin/dashboard.html', 
                         districts=DISTRICTS, 
                         forms=FORM_CONFIGS,
                         recent_entries=recent_entries,
                         recent_uploads=recent_uploads,
                         stats={
                             'total_districts': total_districts,
                             'total_forms': total_forms,
                             'today_entries': today_entries
                         })

@admin_bp.route('/district/<district_name>')
@login_required
def district_view(district_name):
    if current_user.user_type != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    # Get all DSR entries for this district
    entries = DSREntry.query.filter_by(district_name=district_name).order_by(DSREntry.date.desc()).all()
    
    # Group entries by date
    entries_by_date = {}
    for entry in entries:
        date_str = entry.date.strftime('%Y-%m-%d')
        if date_str not in entries_by_date:
            entries_by_date[date_str] = []
        entries_by_date[date_str].append(entry)
    
    return render_template('admin/district_view.html', 
                         district_name=district_name,
                         entries_by_date=entries_by_date,
                         forms=FORM_CONFIGS)

@admin_bp.route('/form/<form_type>')
@login_required
def form_view(form_type):
    if current_user.user_type != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    if form_type not in FORM_CONFIGS:
        flash('Invalid form type', 'error')
        return redirect(url_for('admin.dashboard'))
    
    # Check for date filter parameter
    filter_date = request.args.get('date')
    
    # Get entries for this form type with optional date filter
    query = DSREntry.query.filter_by(form_type=form_type)
    
    if filter_date:
        try:
            # Parse the date string to ensure it's valid
            from datetime import datetime
            parsed_date = datetime.strptime(filter_date, '%Y-%m-%d').date()
            query = query.filter_by(date=parsed_date)
        except ValueError:
            flash('Invalid date format', 'error')
            return redirect(url_for('admin.dashboard'))
    
    entries = query.order_by(DSREntry.date.desc()).all()
    
    # Prepare data for display
    form_data = []
    for entry in entries:
        data = json.loads(entry.data)
        form_data.append({
            'district': entry.district_name,
            'date': entry.date,
            'data': data,
            'id': entry.id
        })
    
    return render_template('admin/form_view.html', 
                         form_type=form_type,
                         form_config=FORM_CONFIGS[form_type],
                         form_data=form_data,
                         filter_date=filter_date)

@admin_bp.route('/search')
@login_required
def search():
    if current_user.user_type != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    
    search_date = request.args.get('date')
    district_name = request.args.get('district')
    
    if not search_date or not district_name:
        return jsonify({'error': 'Date and district required'}), 400
    
    try:
        search_date_obj = datetime.strptime(search_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400
    
    entries = DSREntry.query.filter_by(date=search_date_obj, district_name=district_name).all()
    
    result = []
    for entry in entries:
        data = json.loads(entry.data)
        result.append({
            'form_type': entry.form_type,
            'form_name': FORM_CONFIGS.get(entry.form_type, {}).get('name', entry.form_type),
            'data': data,
            'id': entry.id
        })
    
    return jsonify(result)

@admin_bp.route('/download_dsr/<district_name>/<date_str>')
@login_required
def download_dsr(district_name, date_str):
    if current_user.user_type != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    try:
        search_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date format', 'error')
        return redirect(url_for('admin.dashboard'))
    
    entries = DSREntry.query.filter_by(date=search_date, district_name=district_name).all()
    
    if not entries:
        flash('No data found for the selected date', 'error')
        return redirect(url_for('admin.dashboard'))
    
    # Create Excel file
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"DSR_{district_name}_{date_str}"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Add title
    ws['A1'] = f"Daily Status Report - {district_name}"
    ws['A2'] = f"Date: {date_str}"
    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')
    
    row = 4
    for entry in entries:
        form_config = FORM_CONFIGS.get(entry.form_type, {})
        form_name = form_config.get('name', entry.form_type)
        
        # Form header
        ws[f'A{row}'] = form_name
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Form data
        data = json.loads(entry.data)
        fields = form_config.get('fields', [])
        
        for field in fields:
            field_name = field['name']
            field_label = field['label']
            value = data.get(field_name, '')
            
            ws[f'A{row}'] = field_label
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].border = thin_border
            row += 1
        
        row += 1  # Empty row between forms
    
    wb.save(output)
    output.seek(0)
    
    filename = f"DSR_{district_name}_{date_str}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@admin_bp.route('/uploads')
@login_required
def uploads_view():
    if current_user.user_type != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    # Get filter parameters
    page = request.args.get('page', 1, type=int)
    date_filter = request.args.get('date')
    type_filter = request.args.get('type')
    
    # Build query
    query = ControlRoomUpload.query
    
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            query = query.filter(ControlRoomUpload.date == filter_date)
        except ValueError:
            flash('Invalid date format', 'error')
    
    if type_filter:
        query = query.filter(ControlRoomUpload.upload_type == type_filter)
    
    # Apply pagination
    uploads = query.order_by(ControlRoomUpload.uploaded_at.desc()).paginate(
        page=page, per_page=20, error_out=False)
    
    return render_template('admin/uploads.html', uploads=uploads)

@admin_bp.route('/uploads/filter')
@login_required
def filter_uploads():
    if current_user.user_type != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'error': 'Date is required'}), 400
    
    try:
        filter_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400
    
    uploads = ControlRoomUpload.query.filter_by(date=filter_date).order_by(
        ControlRoomUpload.uploaded_at.desc()).all()
    
    upload_list = []
    for upload in uploads:
        upload_list.append({
            'id': upload.id,
            'original_filename': upload.original_filename,
            'upload_type': upload.upload_type,
            'upload_date': upload.date.strftime('%d-%m-%Y'),
            'uploaded_at': upload.uploaded_at.strftime('%d-%m-%Y %H:%M')
        })
    
    return jsonify(upload_list)

@admin_bp.route('/upload/details/<int:upload_id>')
@login_required
def upload_details(upload_id):
    if current_user.user_type != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    
    upload = ControlRoomUpload.query.get_or_404(upload_id)
    
    # Get file size if file exists
    file_size = None
    try:
        import os
        if os.path.exists(upload.file_path):
            size_bytes = os.path.getsize(upload.file_path)
            if size_bytes < 1024:
                file_size = f"{size_bytes} B"
            elif size_bytes < 1024 * 1024:
                file_size = f"{size_bytes / 1024:.1f} KB"
            else:
                file_size = f"{size_bytes / (1024 * 1024):.1f} MB"
    except:
        pass
    
    return jsonify({
        'original_filename': upload.original_filename,
        'upload_type': upload.upload_type,
        'upload_date': upload.date.strftime('%d-%m-%Y'),
        'uploaded_at': upload.uploaded_at.strftime('%d-%m-%Y %H:%M'),
        'file_size': file_size,
        'uploaded_by': 'Control Room'
    })

@admin_bp.route('/download_upload/<int:upload_id>')
@login_required
def download_upload(upload_id):
    if current_user.user_type != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    upload = ControlRoomUpload.query.get_or_404(upload_id)
    
    try:
        return send_file(
            upload.file_path,
            as_attachment=True,
            download_name=upload.original_filename
        )
    except FileNotFoundError:
        flash('File not found on server', 'error')
        return redirect(url_for('admin.dashboard'))

@admin_bp.route('/view_upload/<int:upload_id>')
@login_required
def view_upload(upload_id):
    if current_user.user_type != 'admin':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    upload = ControlRoomUpload.query.get_or_404(upload_id)
    
    try:
        # Serve file inline (view in browser) instead of download
        return send_file(
            upload.file_path,
            as_attachment=False,  # This makes it view instead of download
            download_name=upload.original_filename
        )
    except FileNotFoundError:
        flash('File not found on server', 'error')
        return redirect(url_for('admin.dashboard'))

# District routes
@district_bp.route('/dashboard')
@login_required
def dashboard():
    if current_user.user_type not in ['district', 'controlroom']:
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    if current_user.user_type == 'controlroom':
        return render_template('district/controlroom_dashboard.html')
    
    # Get recent entries for this district
    recent_entries = DSREntry.query.filter_by(district_name=current_user.district_name)\
                                  .order_by(DSREntry.updated_at.desc())\
                                  .limit(10).all()
    
    return render_template('district/dashboard.html', 
                         forms=FORM_CONFIGS,
                         recent_entries=recent_entries)

@district_bp.route('/controlroom_dashboard')
@login_required
def controlroom_dashboard():
    if current_user.user_type != 'controlroom':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    # Get recent uploads for current user
    recent_uploads = ControlRoomUpload.query.filter_by(user_id=current_user.id).order_by(ControlRoomUpload.uploaded_at.desc()).limit(10).all()
    
    return render_template('district/controlroom_dashboard.html', recent_uploads=recent_uploads)

@district_bp.route('/form/<form_type>', methods=['GET', 'POST'])
@login_required
def form_entry(form_type):
    if current_user.user_type != 'district':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    if form_type not in FORM_CONFIGS:
        flash('Invalid form type', 'error')
        return redirect(url_for('district.dashboard'))
    
    form_config = FORM_CONFIGS[form_type]
    
    if request.method == 'POST':
        entry_date = request.form.get('entry_date')
        if not entry_date:
            flash('Date is required', 'error')
            return render_template('district/form_entry.html', 
                                 form_type=form_type, 
                                 form_config=form_config)
        
        try:
            entry_date_obj = datetime.strptime(entry_date, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid date format', 'error')
            return render_template('district/form_entry.html', 
                                 form_type=form_type, 
                                 form_config=form_config)
        
        # Collect form data
        form_data = {}
        for field in form_config['fields']:
            field_name = field['name']
            form_data[field_name] = request.form.get(field_name, '')
        
        # Check if updating existing entry
        entry_id = request.form.get('entry_id')
        
        if entry_id:
            # Update existing entry
            existing_entry = DSREntry.query.filter_by(
                id=entry_id,
                district_name=current_user.district_name,
                form_type=form_type
            ).first()
            
            if existing_entry:
                existing_entry.data = json.dumps(form_data)
                existing_entry.updated_at = datetime.utcnow()
                db.session.commit()
                flash('Entry updated successfully', 'success')
            else:
                flash('Entry not found', 'error')
        else:
            # Create new entry (multiple entries allowed per date)
            new_entry = DSREntry(
                district_name=current_user.district_name,
                form_type=form_type,
                date=entry_date_obj,
                data=json.dumps(form_data),
                user_id=current_user.id
            )
            db.session.add(new_entry)
            db.session.commit()
            flash('New entry added successfully', 'success')
        
        # Handle AJAX requests (check for XMLHttpRequest header or JSON response preference)
        if (request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 
            'application/json' in request.headers.get('Accept', '') or
            request.form.get('ajax_request') == 'true'):
            return jsonify({'success': True, 'message': 'Entry saved successfully'})
        
        return redirect(url_for('district.dashboard'))
    
    # For GET request, get all existing entries for today
    today = date.today()
    existing_entries = DSREntry.query.filter_by(
        district_name=current_user.district_name,
        form_type=form_type,
        date=today
    ).order_by(DSREntry.created_at.desc()).all()
    
    # Add parsed data to entries for display
    for entry in existing_entries:
        entry.data_dict = json.loads(entry.data)
    
    return render_template('district/form_entry.html', 
                         form_type=form_type, 
                         form_config=form_config,
                         existing_entries=existing_entries,
                         today=today.strftime('%Y-%m-%d'))

@district_bp.route('/form/<form_type>/edit/<int:entry_id>')
@login_required
def edit_entry(form_type, entry_id):
    if current_user.user_type != 'district':
        return jsonify({'success': False, 'message': 'Access denied'})
    
    entry = DSREntry.query.filter_by(
        id=entry_id,
        district_name=current_user.district_name,
        form_type=form_type
    ).first()
    
    if not entry:
        return jsonify({'success': False, 'message': 'Entry not found'})
    
    return jsonify({
        'success': True,
        'entry_data': json.loads(entry.data)
    })

@district_bp.route('/form/<form_type>/delete/<int:entry_id>', methods=['DELETE'])
@login_required
def delete_entry(form_type, entry_id):
    if current_user.user_type != 'district':
        return jsonify({'success': False, 'message': 'Access denied'})
    
    entry = DSREntry.query.filter_by(
        id=entry_id,
        district_name=current_user.district_name,
        form_type=form_type
    ).first()
    
    if not entry:
        return jsonify({'success': False, 'message': 'Entry not found'})
    
    db.session.delete(entry)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Entry deleted successfully'})

@district_bp.route('/upload', methods=['GET', 'POST'])
@login_required
def upload_file():
    if current_user.user_type != 'controlroom':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    if request.method == 'POST':
        upload_date = request.form.get('upload_date')
        upload_type = request.form.get('upload_type')
        file = request.files.get('file')
        
        if not all([upload_date, upload_type, file]):
            flash('All fields are required', 'error')
            return render_template('district/upload.html')
        
        if file.filename == '':
            flash('No file selected', 'error')
            return render_template('district/upload.html')
        
        try:
            upload_date_obj = datetime.strptime(upload_date, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid date format', 'error')
            return render_template('district/upload.html')
        
        # Secure filename
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        new_filename = f"{upload_type}_{timestamp}_{filename}"
        
        # Ensure upload directory exists
        from flask import current_app
        upload_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'controlroom')
        os.makedirs(upload_path, exist_ok=True)
        
        file_path = os.path.join(upload_path, new_filename)
        file.save(file_path)
        
        # Save to database
        upload_record = ControlRoomUpload(
            date=upload_date_obj,
            upload_type=upload_type,
            filename=new_filename,
            original_filename=filename,
            file_path=file_path,
            user_id=current_user.id
        )
        
        db.session.add(upload_record)
        db.session.commit()
        
        flash('File uploaded successfully', 'success')
        return redirect(url_for('district.controlroom_dashboard'))
    
    return render_template('district/upload.html')

@district_bp.route('/download_upload/<int:upload_id>')
@login_required
def download_upload(upload_id):
    if current_user.user_type != 'controlroom':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    upload = ControlRoomUpload.query.get_or_404(upload_id)
    
    # Check if file belongs to current user
    if upload.user_id != current_user.id:
        flash('Access denied', 'error')
        return redirect(url_for('district.controlroom_dashboard'))
    
    try:
        return send_file(
            upload.file_path,
            as_attachment=True,
            download_name=upload.original_filename
        )
    except FileNotFoundError:
        flash('File not found', 'error')
        return redirect(url_for('district.controlroom_dashboard'))

@district_bp.route('/view_upload/<int:upload_id>')
@login_required
def view_upload(upload_id):
    if current_user.user_type != 'controlroom':
        flash('Access denied', 'error')
        return redirect(url_for('main.index'))
    
    upload = ControlRoomUpload.query.get_or_404(upload_id)
    
    # Check if file belongs to current user
    if upload.user_id != current_user.id:
        flash('Access denied', 'error')
        return redirect(url_for('district.controlroom_dashboard'))
    
    try:
        # Serve file inline (view in browser) instead of download
        return send_file(
            upload.file_path,
            as_attachment=False,  # This makes it view instead of download
            download_name=upload.original_filename
        )
    except FileNotFoundError:
        flash('File not found', 'error')
        return redirect(url_for('district.controlroom_dashboard'))

# Profile and Password Management Routes
@auth_bp.route('/profile')
@login_required
def profile():
    return render_template('profile.html', user=current_user)

@auth_bp.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        # Validate current password
        if not current_user.check_password(current_password):
            flash('Current password is incorrect', 'error')
            return render_template('change_password.html')
        
        # Validate new password
        if len(new_password) < 6:
            flash('New password must be at least 6 characters long', 'error')
            return render_template('change_password.html')
        
        if new_password != confirm_password:
            flash('New password and confirmation do not match', 'error')
            return render_template('change_password.html')
        
        # Update password
        current_user.set_password(new_password)
        db.session.commit()
        
        flash('Password changed successfully!', 'success')
        
        # Redirect based on user type
        if current_user.user_type == 'admin':
            return redirect(url_for('admin.dashboard'))
        elif current_user.user_type == 'district':
            return redirect(url_for('district.dashboard'))
        elif current_user.user_type == 'controlroom':
            return redirect(url_for('district.controlroom_dashboard'))
    
    return render_template('change_password.html')